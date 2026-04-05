import streamlit as st
import pandas as pd
import io
import random
from datetime import timedelta
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Taryel Ultimate Logic", layout="wide")

st.title("🚗 Taryel-Logik: Realistische Schichten & Standorte")
st.markdown("Dieses Skript bereinigt \N, simuliert Standorte und glättet die Fahrtenfolge.")

with st.sidebar:
    st.header("⚙️ Konfiguration")
    speed_city = st.number_input("Schnitt KM/H für Anfahrt", value=22)
    min_pause = st.slider("Pause zw. Fahrten (Min)", 2, 5, 3)

uploaded_file = st.file_uploader("Rohdaten hochladen", type=["xlsx", "csv"])

FINAL_COLUMNS = [
    "Datum/Uhrzeit Auftragseingang", "Uhrzeit der Auftragsuebermittlung", "Datum der Fahrt", 
    "Fahrtstatus", "Standort des Fahrzeugs bei Auftragsuebermittlung", "Uhrzeit des Fahrtbeginns", 
    "Uhrzeit des Fahrtendes", "Kennzeichen", "Fahrzeugtyp", "Fahrername", 
    "Fahrpreis", "Kilometer", "Abholort", "Zielort"
]

if uploaded_file:
    try:
        # Einlesen
        df = pd.read_csv(uploaded_file, sep=None, engine='python') if uploaded_file.name.endswith('.csv') else pd.read_excel(uploaded_file)
        df.columns = [str(c).strip() for c in df.columns]

        # 1. Nur abgeschlossene Fahrten
        df = df[df["Fahrtstatus"].str.lower() == "abgeschlossen"]
        
        # 2. Zeit-Parsing & \N Korrektur
        date_cols = ["Uhrzeit des Fahrtbeginns", "Uhrzeit des Fahrtendes", "Datum/Uhrzeit Auftragseingang", "Uhrzeit der Auftragsuebermittlung"]
        for col in date_cols:
            df[col] = pd.to_datetime(df[col], errors='coerce')
        
        df = df.dropna(subset=["Uhrzeit des Fahrtbeginns", "Kennzeichen"])

        output = io.BytesIO()
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df['Tag_Key'] = df['Uhrzeit des Fahrtbeginns'].dt.date
            
            for (tag, kennzeichen, fahrer), group in df.groupby(['Tag_Key', 'Kennzeichen', 'Fahrername']):
                group = group.sort_values("Uhrzeit des Fahrtbeginns")
                final_rows = []
                
                # Wir merken uns die letzte Position (Koordinaten)
                # Falls leer, nehmen wir Kölner Zentrum als Start
                last_coords = "50.9375 6.9603" 

                for i in range(len(group)):
                    row = group.iloc[i].to_dict()
                    
                    # Logik: Ist diese Fahrt sinnvoll? 
                    # Wenn die Fahrt zu weit weg vom letzten Ziel ist oder eine Marzellenstraße-Lücke reißt:
                    # (Hier wird im Taryel-Stil die Zeit-Kette wichtiger gewichtet als der reale Uber-Zeitstempel)
                    
                    if i == 0:
                        # Erste Fahrt: Standort simulieren falls \N
                        if pd.isna(row["Standort des Fahrzeugs bei Auftragsuebermittlung"]) or row["Standort des Fahrzeugs bei Auftragsuebermittlung"] == "\\N":
                            row["Standort des Fahrzeugs bei Auftragsuebermittlung"] = last_coords
                        final_rows.append(row)
                        continue

                    # Vorherige Daten
                    prev = final_rows[-1]
                    prev_ende = prev["Uhrzeit des Fahrtendes"]
                    
                    # 1. STANDORT-LOGIK: Standort bei Übermittlung = Letzter Zielort
                    # Wir simulieren Koordinaten, die nahe am letzten Zielort liegen
                    if not pd.isna(prev.get("Standort des Fahrzeugs bei Auftragsuebermittlung")):
                        # Wir variieren die letzte Koordinate minimal, damit es echt aussieht
                        lat, lon = map(float, str(prev["Standort des Fahrzeugs bei Auftragsuebermittlung"]).split())
                        last_coords = f"{round(lat + random.uniform(-0.001, 0.001), 6)} {round(lon + random.uniform(-0.001, 0.001), 6)}"
                    
                    row["Standort des Fahrzeugs bei Auftragsuebermittlung"] = last_coords

                    # 2. ZEIT-KETTE: Lückenloses "Andocken"
                    wait_time = random.randint(min_pause, min_pause + 3)
                    new_auftrag = prev_ende + timedelta(minutes=wait_time)
                    new_start = new_auftrag + timedelta(minutes=random.randint(2, 5))
                    
                    # Dauer der Originalfahrt erhalten
                    original_duration = row["Uhrzeit des Fahrtendes"] - row["Uhrzeit des Fahrtbeginns"]
                    
                    row["Datum/Uhrzeit Auftragseingang"] = new_auftrag - timedelta(seconds=random.randint(30, 90))
                    row["Uhrzeit der Auftragsuebermittlung"] = new_auftrag
                    row["Uhrzeit des Fahrtbeginns"] = new_start
                    row["Uhrzeit des Fahrtendes"] = new_start + original_duration
                    
                    # 3. KILOMETER-LOGIK
                    # Zeitdifferenz zur Originalzeit wird als "Anfahrt" berechnet
                    time_gap_to_orig = (group.iloc[i]["Uhrzeit des Fahrtbeginns"] - new_start).total_seconds() / 60
                    if time_gap_to_orig > 0:
                        bonus_km = round(time_gap_to_orig * (speed_city / 60), 2)
                        row["Kilometer"] = round(float(row.get("Kilometer", 0)) + bonus_km, 2)

                    row["_CORRECTED"] = True
                    final_rows.append(row)

                # Export-Vorbereitung
                res_df = pd.DataFrame(final_rows)
                for c in date_cols:
                    res_df[c] = pd.to_datetime(res_df[c]).dt.strftime('%Y-%m-%d %H:%M:%S')

                sheet_name = f"{tag}_{fahrer[:10]}".replace("/", "")[:31]
                res_df[FINAL_COLUMNS].to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Färbung
                ws = writer.sheets[sheet_name]
                for idx, r_data in enumerate(final_rows, start=2):
                    if r_data.get("_CORRECTED"):
                        for c_idx in range(1, len(FINAL_COLUMNS) + 1):
                            ws.cell(row=idx, column=c_idx).fill = orange_fill

        st.success("✅ Fertig! Standorte sind korrigiert, Zeitketten sind lückenlos.")
        st.download_button("Perfektionierte Taryel-Liste laden", data=output.getvalue(), file_name="Uber_Taryel_Final_Safe.xlsx")
            
    except Exception as e:
        st.error(f"Fehler: {e}")
